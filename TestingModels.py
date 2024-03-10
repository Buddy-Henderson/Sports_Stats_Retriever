from main_Library import *
from prediction_Library import *
from calculation_Library import *

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

from datetime import datetime

# Testing Calls
"""
    - To test NBA Pred Points model: (team1, team2)
    - To test NBA Pred Threes model: (team1, team2)
    - To test NBA Pred Rebounds model: (team1, team2)
    
"""




def test_NBA_PredPoints(teamName, vs_teamName):
    
    # Connect to your MySQL database
    db_host = '127.0.0.1'  # Replace with your database host
    db_user = 'root'  # Replace with your database username
    db_password = 'root'  # Replace with your database password
    db_name = 'nba_stats'  # Replace with your database name

    conn = mysql.connector.connect(
        host=db_host,
        user=db_user,
        password=db_password,
        database=db_name
    )
    
    cursor = conn.cursor()
    
    current_date = datetime.now().date()
    
    
    if teamName == "Boston":
        
        team1_short = "bos"
        
    
    elif teamName == "Brooklyn":
        
        team1_short = "bkn"
        team_short = "brooklyn-nets"
        
    elif teamName == "New York":
        
        team1_short = "ny"
        
        
    elif teamName == "Philadelphia":
        
        team1_short = "phi"
        
        
    elif teamName == "Toronto":
        
        team1_short = "tor"
        

    elif teamName == "Golden State":
        
        team1_short = "gs"
        
        
    elif teamName == "LA Clippers":
        
        team1_short = "lac"
        
        
    elif teamName == "LA Lakers":
        
        team1_short = "lal"
        
        
    elif teamName == "Phoenix": 
        
        team1_short = "phx"  
            
        
    elif teamName == "Sacramento":
    
        team1_short = "sac"
        
        
    elif teamName == "Chicago":
        
        team1_short = "chi"
        
        
    elif teamName == "Cleveland":
        
        team1_short = "cle"
        
        
    elif teamName == "Detroit":
        
        team1_short = "det"
        
        
    elif teamName == "Indiana":
        
        team1_short = "ind"
        
    
    elif teamName == "Milwaukee":
        
        team1_short = "mil"
        
        
    elif teamName == "Dallas":
        
        team1_short = "dal"
        
        
    elif teamName == "Houston":
        
        team1_short = "hou"
        
        
    elif teamName == "Memphis":
        
        team1_short = "mem"
        
        
    elif teamName == "New Orleans":
        
        team1_short = "no"
        
        
    elif teamName == "San Antonio":
        
        team1_short = "sa"
        
        
    elif teamName == "Atlanta":
        
        team1_short = "atl"
        
        
    elif teamName == "Charlotte":
        
        team1_short = "cha"
        
        
    elif teamName == "Miami":
        
        team1_short = "mia"   
            
        
    elif teamName == "Orlando":
        
        team1_short = "orl"
        
    
    elif teamName == "Washington":
        
        team1_short = "wsh"
        
        
    elif teamName == "Denver":
        
        team1_short = "den"
        
        
    elif teamName == "Minnesota":
        
        team1_short = "min"
        
        
    elif teamName == "Okla City":
        
        team1_short = "okc"
        
        
    elif teamName == "Portland":
        
        team1_short = "por"
        
        
    elif teamName == "Utah":
        
        team1_short = "utah"
    
    
    
    point_Prediction = []
    
    query = f"SELECT Name FROM {team1_short}_roster WHERE String = %s"
    cursor.execute(query, (1,))
    result = cursor.fetchall()
    
    startingPlayers = [row[0] for row in result]
    
    for player in startingPlayers:
        
        playerHasStats = checkNBA_PlayerHasStats(player)
        
        if playerHasStats:
            pred_Points =getNBAPred(player, "Offense",vs_teamName, "Player_Pred_Points")
            pred_Points = pred_Points[0]
            
            point_Prediction.append(pred_Points)
    
    
    # Load the existing workbook
    try:
        wb = load_workbook(filename=f"C:\\Users\\buddy\\OneDrive\\Desktop\\Miami\\Senior Capstone\\Model Testing\\NBA\Points Predictions\\predPoints_Model_{current_date}.xlsx")
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        
    ws = wb.active
    
    # Find the last row with data in column A
    last_row = ws.max_row
    
    # Set up headers and subheaders starting from the last row + 2
    ws.cell(row=last_row + 2, column=1, value=str(current_date))
    ws.merge_cells(start_row=last_row + 2, start_column=1, end_row=last_row + 2, end_column=2)  # Merge cells for the header
    
    ws.cell(row=last_row + 3, column=1, value='Pred Player Points')
    ws.merge_cells(start_row=last_row + 3, start_column=1, end_row=last_row + 3, end_column=2)  # Merge cells for the subheader
    
    ws.cell(row=last_row + 4, column=1, value=teamName)
    ws.merge_cells(start_row=last_row + 4, start_column=1, end_row=last_row + 4, end_column=2)
    
    # Populate column A with startingPlayers
    for index, player in enumerate(startingPlayers, start=last_row + 5):  # Start writing from the last row + 5
        ws.cell(row=index, column=1, value=player)
        
    # Populate column B with point_Prediction
    for index, prediction in enumerate(point_Prediction, start=last_row + 5):  # Start writing from the last row + 5
        ws.cell(row=index, column=2, value=round(prediction))
    
    
    wb.save(f"C:\\Users\\buddy\\OneDrive\\Desktop\\Miami\\Senior Capstone\\Model Testing\\NBA\Points Predictions\\predPoints_Model_{current_date}.xlsx")

def test_NBA_PredPoints2(teamName, vs_teamName):
    
    # Connect to your MySQL database
    db_host = '127.0.0.1'  # Replace with your database host
    db_user = 'root'  # Replace with your database username
    db_password = 'root'  # Replace with your database password
    db_name = 'nba_stats'  # Replace with your database name

    conn = mysql.connector.connect(
        host=db_host,
        user=db_user,
        password=db_password,
        database=db_name
    )
    
    cursor = conn.cursor()
    
    current_date = datetime.now().date()
    
    
    if teamName == "Boston":
        
        team1_short = "bos"
        
    
    elif teamName == "Brooklyn":
        
        team1_short = "bkn"
        team_short = "brooklyn-nets"
        
    elif teamName == "New York":
        
        team1_short = "ny"
        
        
    elif teamName == "Philadelphia":
        
        team1_short = "phi"
        
        
    elif teamName == "Toronto":
        
        team1_short = "tor"
        

    elif teamName == "Golden State":
        
        team1_short = "gs"
        
        
    elif teamName == "LA Clippers":
        
        team1_short = "lac"
        
        
    elif teamName == "LA Lakers":
        
        team1_short = "lal"
        
        
    elif teamName == "Phoenix": 
        
        team1_short = "phx"  
            
        
    elif teamName == "Sacramento":
    
        team1_short = "sac"
        
        
    elif teamName == "Chicago":
        
        team1_short = "chi"
        
        
    elif teamName == "Cleveland":
        
        team1_short = "cle"
        
        
    elif teamName == "Detroit":
        
        team1_short = "det"
        
        
    elif teamName == "Indiana":
        
        team1_short = "ind"
        
    
    elif teamName == "Milwaukee":
        
        team1_short = "mil"
        
        
    elif teamName == "Dallas":
        
        team1_short = "dal"
        
        
    elif teamName == "Houston":
        
        team1_short = "hou"
        
        
    elif teamName == "Memphis":
        
        team1_short = "mem"
        
        
    elif teamName == "New Orleans":
        
        team1_short = "no"
        
        
    elif teamName == "San Antonio":
        
        team1_short = "sa"
        
        
    elif teamName == "Atlanta":
        
        team1_short = "atl"
        
        
    elif teamName == "Charlotte":
        
        team1_short = "cha"
        
        
    elif teamName == "Miami":
        
        team1_short = "mia"   
            
        
    elif teamName == "Orlando":
        
        team1_short = "orl"
        
    
    elif teamName == "Washington":
        
        team1_short = "wsh"
        
        
    elif teamName == "Denver":
        
        team1_short = "den"
        
        
    elif teamName == "Minnesota":
        
        team1_short = "min"
        
        
    elif teamName == "Okla City":
        
        team1_short = "okc"
        
        
    elif teamName == "Portland":
        
        team1_short = "por"
        
        
    elif teamName == "Utah":
        
        team1_short = "utah"
    
    
    
    point_Prediction = []
    
    query = f"SELECT Name FROM {team1_short}_roster WHERE String = %s"
    cursor.execute(query, (1,))
    result = cursor.fetchall()
    
    startingPlayers = [row[0] for row in result]
    
    for player in startingPlayers:
        
        playerHasStats = checkNBA_PlayerHasStats(player)
        
        if playerHasStats:
            pred_Points =getNBAPred(player, "Offense",vs_teamName, "Player_Pred_Points2")
            pred_Points = pred_Points[0]
            
            point_Prediction.append(pred_Points)
    
    
    # Load the existing workbook
    try:
        wb = load_workbook(filename=f"C:\\Users\\buddy\\OneDrive\\Desktop\\Miami\\Senior Capstone\\Model Testing\\NBA\Points Predictions\\predPoints2_Model_{current_date}.xlsx")
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        
    ws = wb.active
    
    # Find the last row with data in column A
    last_row = ws.max_row
    
    # Set up headers and subheaders starting from the last row + 2
    ws.cell(row=last_row + 2, column=1, value=str(current_date))
    ws.merge_cells(start_row=last_row + 2, start_column=1, end_row=last_row + 2, end_column=2)  # Merge cells for the header
    
    ws.cell(row=last_row + 3, column=1, value='Pred Player Points')
    ws.merge_cells(start_row=last_row + 3, start_column=1, end_row=last_row + 3, end_column=2)  # Merge cells for the subheader
    
    ws.cell(row=last_row + 4, column=1, value=teamName)
    ws.merge_cells(start_row=last_row + 4, start_column=1, end_row=last_row + 4, end_column=2)
    
    # Populate column A with startingPlayers
    for index, player in enumerate(startingPlayers, start=last_row + 5):  # Start writing from the last row + 5
        ws.cell(row=index, column=1, value=player)
        
    # Populate column B with point_Prediction
    for index, prediction in enumerate(point_Prediction, start=last_row + 5):  # Start writing from the last row + 5
        ws.cell(row=index, column=2, value=round(prediction))
    
    
    wb.save(f"C:\\Users\\buddy\\OneDrive\\Desktop\\Miami\\Senior Capstone\\Model Testing\\NBA\Points Predictions\\predPoints2_Model_{current_date}.xlsx")

def test_NBA_PredPoints3(teamName, vs_teamName):
    
    # Connect to your MySQL database
    db_host = '127.0.0.1'  # Replace with your database host
    db_user = 'root'  # Replace with your database username
    db_password = 'root'  # Replace with your database password
    db_name = 'nba_stats'  # Replace with your database name

    conn = mysql.connector.connect(
        host=db_host,
        user=db_user,
        password=db_password,
        database=db_name
    )
    
    cursor = conn.cursor()
    
    current_date = datetime.now().date()
    
    
    if teamName == "Boston":
        
        team1_short = "bos"
        
    
    elif teamName == "Brooklyn":
        
        team1_short = "bkn"
        team_short = "brooklyn-nets"
        
    elif teamName == "New York":
        
        team1_short = "ny"
        
        
    elif teamName == "Philadelphia":
        
        team1_short = "phi"
        
        
    elif teamName == "Toronto":
        
        team1_short = "tor"
        

    elif teamName == "Golden State":
        
        team1_short = "gs"
        
        
    elif teamName == "LA Clippers":
        
        team1_short = "lac"
        
        
    elif teamName == "LA Lakers":
        
        team1_short = "lal"
        
        
    elif teamName == "Phoenix": 
        
        team1_short = "phx"  
            
        
    elif teamName == "Sacramento":
    
        team1_short = "sac"
        
        
    elif teamName == "Chicago":
        
        team1_short = "chi"
        
        
    elif teamName == "Cleveland":
        
        team1_short = "cle"
        
        
    elif teamName == "Detroit":
        
        team1_short = "det"
        
        
    elif teamName == "Indiana":
        
        team1_short = "ind"
        
    
    elif teamName == "Milwaukee":
        
        team1_short = "mil"
        
        
    elif teamName == "Dallas":
        
        team1_short = "dal"
        
        
    elif teamName == "Houston":
        
        team1_short = "hou"
        
        
    elif teamName == "Memphis":
        
        team1_short = "mem"
        
        
    elif teamName == "New Orleans":
        
        team1_short = "no"
        
        
    elif teamName == "San Antonio":
        
        team1_short = "sa"
        
        
    elif teamName == "Atlanta":
        
        team1_short = "atl"
        
        
    elif teamName == "Charlotte":
        
        team1_short = "cha"
        
        
    elif teamName == "Miami":
        
        team1_short = "mia"   
            
        
    elif teamName == "Orlando":
        
        team1_short = "orl"
        
    
    elif teamName == "Washington":
        
        team1_short = "wsh"
        
        
    elif teamName == "Denver":
        
        team1_short = "den"
        
        
    elif teamName == "Minnesota":
        
        team1_short = "min"
        
        
    elif teamName == "Okla City":
        
        team1_short = "okc"
        
        
    elif teamName == "Portland":
        
        team1_short = "por"
        
        
    elif teamName == "Utah":
        
        team1_short = "utah"
    
    
    
    point_Prediction = []
    
    query = f"SELECT Name FROM {team1_short}_roster WHERE String = %s"
    cursor.execute(query, (1,))
    result = cursor.fetchall()
    
    startingPlayers = [row[0] for row in result]
    
    for player in startingPlayers:
        
        playerHasStats = checkNBA_PlayerHasStats(player)
        
        if playerHasStats:
            pred_Points =getNBAPred(player, "Offense",vs_teamName, "Player_Pred_Points3")
            pred_Points = pred_Points[0]
            
            point_Prediction.append(pred_Points)
    
    
    # Load the existing workbook
    try:
        wb = load_workbook(filename=f"C:\\Users\\buddy\\OneDrive\\Desktop\\Miami\\Senior Capstone\\Model Testing\\NBA\Points Predictions\\predPoints3_Model_{current_date}.xlsx")
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        
    ws = wb.active
    
    # Find the last row with data in column A
    last_row = ws.max_row
    
    # Set up headers and subheaders starting from the last row + 2
    ws.cell(row=last_row + 2, column=1, value=str(current_date))
    ws.merge_cells(start_row=last_row + 2, start_column=1, end_row=last_row + 2, end_column=2)  # Merge cells for the header
    
    ws.cell(row=last_row + 3, column=1, value='Pred Player Points')
    ws.merge_cells(start_row=last_row + 3, start_column=1, end_row=last_row + 3, end_column=2)  # Merge cells for the subheader
    
    ws.cell(row=last_row + 4, column=1, value=teamName)
    ws.merge_cells(start_row=last_row + 4, start_column=1, end_row=last_row + 4, end_column=2)
    
    # Populate column A with startingPlayers
    for index, player in enumerate(startingPlayers, start=last_row + 5):  # Start writing from the last row + 5
        ws.cell(row=index, column=1, value=player)
        
    # Populate column B with point_Prediction
    for index, prediction in enumerate(point_Prediction, start=last_row + 5):  # Start writing from the last row + 5
        ws.cell(row=index, column=2, value=round(prediction))
    
    
    wb.save(f"C:\\Users\\buddy\\OneDrive\\Desktop\\Miami\\Senior Capstone\\Model Testing\\NBA\Points Predictions\\predPoints3_Model_{current_date}.xlsx")

def test_NBA_PredThrees(teamName, vs_teamName):
    
    # Connect to your MySQL database
    db_host = '127.0.0.1'  # Replace with your database host
    db_user = 'root'  # Replace with your database username
    db_password = 'root'  # Replace with your database password
    db_name = 'nba_stats'  # Replace with your database name

    conn = mysql.connector.connect(
        host=db_host,
        user=db_user,
        password=db_password,
        database=db_name
    )
    
    cursor = conn.cursor()
    
    current_date = datetime.now().date()
    
    
    if teamName == "Boston":
        
        team1_short = "bos"
        
    
    elif teamName == "Brooklyn":
        
        team1_short = "bkn"
        
        
    elif teamName == "New York":
        
        team1_short = "ny"
        
        
    elif teamName == "Philadelphia":
        
        team1_short = "phi"
        
        
    elif teamName == "Toronto":
        
        team1_short = "tor"
        

    elif teamName == "Golden State":
        
        team1_short = "gs"
        
        
    elif teamName == "LA Clippers":
        
        team1_short = "lac"
        
        
    elif teamName == "LA Lakers":
        
        team1_short = "lal"
        
        
    elif teamName == "Phoenix": 
        
        team1_short = "phx"  
            
        
    elif teamName == "Sacramento":
    
        team1_short = "sac"
        
        
    elif teamName == "Chicago":
        
        team1_short = "chi"
        
        
    elif teamName == "Cleveland":
        
        team1_short = "cle"
        
        
    elif teamName == "Detroit":
        
        team1_short = "det"
        
        
    elif teamName == "Indiana":
        
        team1_short = "ind"
        
    
    elif teamName == "Milwaukee":
        
        team1_short = "mil"
        
        
    elif teamName == "Dallas":
        
        team1_short = "dal"
        
        
    elif teamName == "Houston":
        
        team1_short = "hou"
        
        
    elif teamName == "Memphis":
        
        team1_short = "mem"
        
        
    elif teamName == "New Orleans":
        
        team1_short = "no"
        
        
    elif teamName == "San Antonio":
        
        team1_short = "sa"
        
        
    elif teamName == "Atlanta":
        
        team1_short = "atl"
        
        
    elif teamName == "Charlotte":
        
        team1_short = "cha"
        
        
    elif teamName == "Miami":
        
        team1_short = "mia"   
            
        
    elif teamName == "Orlando":
        
        team1_short = "orl"
        
    
    elif teamName == "Washington":
        
        team1_short = "wsh"
        
        
    elif teamName == "Denver":
        
        team1_short = "den"
        
        
    elif teamName == "Minnesota":
        
        team1_short = "min"
        
        
    elif teamName == "Okla City":
        
        team1_short = "okc"
        
        
    elif teamName == "Portland":
        
        team1_short = "por"
        
        
    elif teamName == "Utah":
        
        team1_short = "utah"
    
    
    
    Threes_Prediction = []
    
    query = f"SELECT Name FROM {team1_short}_roster WHERE String = %s"
    cursor.execute(query, (1,))
    result = cursor.fetchall()
    
    startingPlayers = [row[0] for row in result]
    
    for player in startingPlayers:
        
        playerHasStats = checkNBA_PlayerHasStats(player)
        
        if playerHasStats:
            pred_Threes =getNBAPred(player, "Offense",vs_teamName, "Player_Pred_Threes")
            pred_Threes = pred_Threes[0]
            
            Threes_Prediction.append(pred_Threes)
        
    # Load the existing workbook
    try:
        wb = load_workbook(filename=f"C:\\Users\\buddy\\OneDrive\\Desktop\\Miami\\Senior Capstone\\Model Testing\\NBA\Points Predictions\\predThrees_Model_{current_date}.xlsx")
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        
    ws = wb.active
    
    # Find the last row with data in column A
    last_row = ws.max_row
    
    # Set up headers and subheaders starting from the last row + 2
    ws.cell(row=last_row + 2, column=1, value=str(current_date))
    ws.merge_cells(start_row=last_row + 2, start_column=1, end_row=last_row + 2, end_column=2)  # Merge cells for the header
    
    ws.cell(row=last_row + 3, column=1, value='Pred Player Threes')
    ws.merge_cells(start_row=last_row + 3, start_column=1, end_row=last_row + 3, end_column=2)  # Merge cells for the subheader
    
    ws.cell(row=last_row + 4, column=1, value=teamName)
    ws.merge_cells(start_row=last_row + 4, start_column=1, end_row=last_row + 4, end_column=2)
    
    # Populate column A with startingPlayers
    for index, player in enumerate(startingPlayers, start=last_row + 5):  # Start writing from the last row + 5
        ws.cell(row=index, column=1, value=player)
        
    # Populate column B with point_Prediction
    for index, prediction in enumerate(Threes_Prediction, start=last_row + 5):  # Start writing from the last row + 5
        ws.cell(row=index, column=2, value=round(prediction))
    
    
    wb.save(f"C:\\Users\\buddy\\OneDrive\\Desktop\\Miami\\Senior Capstone\\Model Testing\\NBA\Points Predictions\\predThrees_Model_{current_date}.xlsx")

def test_NBA_PredRebounds(teamName, vs_teamName):

    # Connect to your MySQL database
    db_host = '127.0.0.1'  # Replace with your database host
    db_user = 'root'  # Replace with your database username
    db_password = 'root'  # Replace with your database password
    db_name = 'nba_stats'  # Replace with your database name

    conn = mysql.connector.connect(
        host=db_host,
        user=db_user,
        password=db_password,
        database=db_name
    )
    
    cursor = conn.cursor()
    
    current_date = datetime.now().date()
    
    
    if teamName == "Boston":
        
        team1_short = "bos"
        
    
    elif teamName == "Brooklyn":
        
        team1_short = "bkn"
        
        
    elif teamName == "New York":
        
        team1_short = "ny"
        
        
    elif teamName == "Philadelphia":
        
        team1_short = "phi"
        
        
    elif teamName == "Toronto":
        
        team1_short = "tor"
        

    elif teamName == "Golden State":
        
        team1_short = "gs"
        
        
    elif teamName == "LA Clippers":
        
        team1_short = "lac"
        
        
    elif teamName == "LA Lakers":
        
        team1_short = "lal"
        
        
    elif teamName == "Phoenix": 
        
        team1_short = "phx"  
            
        
    elif teamName == "Sacramento":
    
        team1_short = "sac"
        
        
    elif teamName == "Chicago":
        
        team1_short = "chi"
        
        
    elif teamName == "Cleveland":
        
        team1_short = "cle"
        
        
    elif teamName == "Detroit":
        
        team1_short = "det"
        
        
    elif teamName == "Indiana":
        
        team1_short = "ind"
        
    
    elif teamName == "Milwaukee":
        
        team1_short = "mil"
        
        
    elif teamName == "Dallas":
        
        team1_short = "dal"
        
        
    elif teamName == "Houston":
        
        team1_short = "hou"
        
        
    elif teamName == "Memphis":
        
        team1_short = "mem"
        
        
    elif teamName == "New Orleans":
        
        team1_short = "no"
        
        
    elif teamName == "San Antonio":
        
        team1_short = "sa"
        
        
    elif teamName == "Atlanta":
        
        team1_short = "atl"
        
        
    elif teamName == "Charlotte":
        
        team1_short = "cha"
        
        
    elif teamName == "Miami":
        
        team1_short = "mia"   
            
        
    elif teamName == "Orlando":
        
        team1_short = "orl"
        
    
    elif teamName == "Washington":
        
        team1_short = "wsh"
        
        
    elif teamName == "Denver":
        
        team1_short = "den"
        
        
    elif teamName == "Minnesota":
        
        team1_short = "min"
        
        
    elif teamName == "Okla City":
        
        team1_short = "okc"
        
        
    elif teamName == "Portland":
        
        team1_short = "por"
        
        
    elif teamName == "Utah":
        
        team1_short = "utah"
    
    
    
    rebounds_Prediction = []
    
    query = f"SELECT Name FROM {team1_short}_roster WHERE String = %s"
    cursor.execute(query, (1,))
    result = cursor.fetchall()
    
    startingPlayers = [row[0] for row in result]
    
    for player in startingPlayers:
        
        playerHasStats = checkNBA_PlayerHasStats(player)
        
        if playerHasStats:
            
            pred_Rebounds =getNBAPred(player, "Offense",vs_teamName, "Player_Pred_Rebounds")
            pred_Rebounds = pred_Rebounds[0]
            
            rebounds_Prediction.append(pred_Rebounds)
        
    # Load the existing workbook
    try:
        wb = load_workbook(filename=f"C:\\Users\\buddy\\OneDrive\\Desktop\\Miami\\Senior Capstone\\Model Testing\\NBA\Points Predictions\\predRebounds_Model_{current_date}.xlsx")
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        
    ws = wb.active
    
    # Find the last row with data in column A
    last_row = ws.max_row
    
    # Set up headers and subheaders starting from the last row + 2
    ws.cell(row=last_row + 2, column=1, value=str(current_date))
    ws.merge_cells(start_row=last_row + 2, start_column=1, end_row=last_row + 2, end_column=2)  # Merge cells for the header
    
    ws.cell(row=last_row + 3, column=1, value='Pred Player Rebounds')
    ws.merge_cells(start_row=last_row + 3, start_column=1, end_row=last_row + 3, end_column=2)  # Merge cells for the subheader
    
    ws.cell(row=last_row + 4, column=1, value=teamName)
    ws.merge_cells(start_row=last_row + 4, start_column=1, end_row=last_row + 4, end_column=2)
    
    # Populate column A with startingPlayers
    for index, player in enumerate(startingPlayers, start=last_row + 5):  # Start writing from the last row + 5
        ws.cell(row=index, column=1, value=player)
        
    # Populate column B with point_Prediction
    for index, prediction in enumerate(rebounds_Prediction, start=last_row + 5):  # Start writing from the last row + 5
        ws.cell(row=index, column=2, value=round(prediction))
    
    
    wb.save(f"C:\\Users\\buddy\\OneDrive\\Desktop\\Miami\\Senior Capstone\\Model Testing\\NBA\Points Predictions\\predRebounds_Model_{current_date}.xlsx")

# Call the test models here:
# enter all teams and then who they are playing
teamName = ["Chicago", "LA Clippers", "Dallas", "Detroit", "Brooklyn", "Charlotte", "San Antonio", "Golden State", "Boston", "Phoenix", "Utah", "Denver", "Toronto", "Portland"]
vs_teamName = ["LA Clippers", "Chicago", "Detroit", "Dallas", "Charlotte", "Brooklyn", "Golden State", "San Antonio", "Phoenix", "Boston", "Denver", "Utah", "Portland", "Toronto"]

for team, vs_team in zip(teamName, vs_teamName):

    test_NBA_PredRebounds(team, vs_team)
    